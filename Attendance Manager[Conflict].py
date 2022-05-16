import os
import openpyxl

# C:\Users\Shashank Shekhar\Documents\Attendance.xlsx
# C:\Users\Shashank Shekhar\Documents\ieuofimuav 06_11_2021 12_52_12 PM.xlsx

gmf_input = input('Enter Google Meet Attendance File Path : ')
google_meet_file = openpyxl.load_workbook(f'{gmf_input}')
number_of_rows = google_meet_file.active.max_row
present_students = []
i = 2
for value in range(number_of_rows-1):
    data = google_meet_file.active[f'A{str(i)}'].value
    present_students.append(data.lower())
    i = i + 1
print(f'Students Present : {present_students}')

af_input = input('Enter Attendance File Path : ')
attendance_file = openpyxl.load_workbook(f'{af_input}')
sheet_1 = attendance_file['Sheet1']
number_of_rows2 = sheet_1.max_row
students = []
x = 2
column_number = int(input("Please Enter The Column No. = "))
for value in range(number_of_rows2 - 1):
    data = sheet_1[f'A{str(x)}'].value
    students.append(data.lower())
    if data.lower() in present_students:
        sheet_1.cell(row=x, column=column_number, value='Present')
    else:
        sheet_1.cell(row=x, column=column_number, value='Absent')
    x = x + 1

attendance_file.save(f'{af_input}')
print('File Saved')